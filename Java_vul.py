from datetime import datetime
import csv
import subprocess
import os
from openpyxl import Workbook, load_workbook

def get_timestamped_filename(base_filename):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base_filename}_{timestamp}.xlsx"

def get_timestamped_zip_filename(base_name):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base_name}_{timestamp}.zip"

def zip_directory(hostname, full_path, log_file):
    try:
        base_name = os.path.basename(full_path.rstrip('/'))
        parent_dir = os.path.dirname(full_path.rstrip('/'))
        zip_file_name = get_timestamped_zip_filename(base_name)  # Add timestamp to the zip file name
        zip_file = os.path.join(parent_dir, zip_file_name)
        check_zip_command = f"ssh cloud-user@{hostname} 'command -v zip'"
        check_zip_result = subprocess.run(check_zip_command, shell=True, capture_output=True, text=True)
        if check_zip_result.returncode != 0:
            print(f"zip command not found on {hostname}. Installing zip package...")
            install_zip_command = f"ssh cloud-user@{hostname} 'sudo yum install zip -y'"
            install_zip_result = subprocess.run(install_zip_command, shell=True, capture_output=True, text=True)
            
            if install_zip_result.returncode != 0:
                print(f"Failed to install zip package on {hostname}: {install_zip_result.stderr}")
                return 

        
        zip_command = f"""
        ssh cloud-user@{hostname} 'sudo bash -c "
        if [ -d {full_path} ]; then
            cd {parent_dir} && zip -r {zip_file_name} {base_name};
            echo SUCCESS;
        else
            echo NOT_EXIST;
        fi"'
        """

        print(f"Checking and zipping {base_name} on {hostname} with filename {zip_file_name}...")
        result = subprocess.run(zip_command, shell=True, capture_output=True, text=True)

        if "NOT_EXIST" in result.stdout:
            print(f"Directory {full_path} does not exist on {hostname}. Skipping zipping.")
            return

        if "SUCCESS" in result.stdout:
            print(f"Successfully zipped {base_name} to {zip_file} on {hostname}.")
            if os.path.exists(log_file):
                workbook = load_workbook(log_file)
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["Hostname", "Zipped File Path"])
            sheet = workbook.active
            sheet.append([hostname, zip_file])
            workbook.save(log_file)
        else:
            print(f"Error zipping {base_name} on {hostname}: {result.stderr}")
            print(f"Command output: {result.stdout}")
            print(f"Command error: {result.stderr}")

    except Exception as e:
        print(f"Failed to zip {base_name} on {hostname}. Error: {str(e)}")

def remove_directory(hostname, full_path):
    try:
        remove_command = f"ssh cloud-user@{hostname} 'sudo rm -rf {full_path}'"
        print(f"Removing directory {full_path} on {hostname}...")
        result = subprocess.run(remove_command, shell=True, capture_output=True, text=True)

        if result.returncode == 0:
            print(f"Successfully removed {full_path} on {hostname}.")
        else:
            print(f"Error removing {full_path} on {hostname}: {result.stderr}")

    except Exception as e:
        print(f"Failed to remove {full_path} on {hostname}. Error: {str(e)}")

def reload_oem_agent(hostname):
    try:
        
        reload_command = (
            f"ssh cloud-user@{hostname} 'sudo su - oracle -c \""
            f"/u01/app/oracle/product/OEM135/agent/agent_13.5.0.0.0/bin/emctl stop agent && "
            f"/u01/app/oracle/product/OEM135/agent/agent_13.5.0.0.0/bin/emctl start agent\"'"
        )
        
        print(f"Reloading OEM agent on {hostname}...")
        result = subprocess.run(reload_command, shell=True, capture_output=True, text=True)

        if result.returncode == 0:
            print(f"Successfully reloaded OEM agent on {hostname}.")
        else:
            print(f"Error reloading OEM agent on {hostname}: {result.stderr}")

    except Exception as e:
        print(f"Failed to reload OEM agent on {hostname}. Error: {str(e)}")


def zip_and_remove_directory(hostname, full_path, log_file, archive_path, action, oem_action):
    archive_filename = os.path.basename(archive_path)

    if action.strip().upper() == "UPGRADE" or action.strip().upper() == "JRE_UPGRADE":
        cleanup_command = f"ssh cloud-user@{hostname} 'sudo rm -rf /tmp/jdk /tmp/jre'"
        cleanup_result = subprocess.run(cleanup_command, shell=True, capture_output=True, text=True)
        if cleanup_result.returncode == 0:
            print(f"Successfully removed existing /tmp/jdk or /tmp/jre directories on {hostname}.")
        else:
            print(f"Error removing /tmp/jdk or /tmp/jre on {hostname}: {cleanup_result.stderr}")
            
        zip_directory(hostname, full_path, log_file)
        remove_directory(hostname, full_path)
        print(f"Transferring {archive_path} to /tmp/{archive_filename} on {hostname}...")
        transfer_command = f"scp {archive_path} cloud-user@{hostname}:/tmp/{archive_filename}"
        transfer_result = subprocess.run(transfer_command, shell=True, capture_output=True, text=True)

        if transfer_result.returncode == 0:
            print(f"Successfully transferred {archive_path} to /tmp/{archive_filename} on {hostname}.")
            extract_command = f"ssh cloud-user@{hostname} 'sudo mkdir -p /tmp/{archive_filename.split('.')[0]} && sudo tar -xzvf /tmp/{archive_filename} -C /tmp/{archive_filename.split('.')[0]} && sudo rm -f /tmp/{archive_filename}'"
            extract_result = subprocess.run(extract_command, shell=True, capture_output=True, text=True)

            if extract_result.returncode == 0:
                print(f"Successfully extracted {archive_filename.split('.')[0]} to /tmp/{archive_filename.split('.')[0]} on {hostname}.")
                copy_command = f"ssh cloud-user@{hostname} 'sudo cp -R /tmp/{archive_filename.split('.')[0]}/{archive_filename.split('.')[0]} {os.path.dirname(full_path)}/'"
                copy_result = subprocess.run(copy_command, shell=True, capture_output=True, text=True)

                if copy_result.returncode == 0:
                    print(f"Successfully copied {archive_filename.split('.')[0]} to {os.path.dirname(full_path)} on {hostname}.")
                    chown_command = f"ssh cloud-user@{hostname} 'sudo chown -R oracle:dba {os.path.join(os.path.dirname(full_path), archive_filename.split('.')[0])}'"
                    chown_result = subprocess.run(chown_command, shell=True, capture_output=True, text=True)

                    if chown_result.returncode == 0:
                        print(f"Successfully changed ownership of {archive_filename.split('.')[0]} to oracle:dba on {hostname}.")
                        chmod_command = f"ssh cloud-user@{hostname} 'sudo chmod -R 755 {os.path.join(os.path.dirname(full_path), archive_filename.split('.')[0])}'"
                        chmod_result = subprocess.run(chmod_command, shell=True, capture_output=True, text=True)
                        if chmod_result.returncode == 0:
                            print(f"Successfully set permissions to 755 for {archive_filename.split('.')[0]} on {hostname}.")
                        else:
                            print(f"Error setting permissions to 755 for {archive_filename.split('.')[0]} on {hostname}: {chmod_result.stderr}")
                    else:
                        print(f"Error changing ownership of {archive_filename.split('.')[0]} on {hostname}: {chown_result.stderr}")
        
                    cleanup_command = f"ssh cloud-user@{hostname} 'sudo rm -rf /tmp/{archive_filename.split('.')[0]}'"
                    cleanup_result = subprocess.run(cleanup_command, shell=True, capture_output=True, text=True)

                    if cleanup_result.returncode == 0:
                        print(f"Successfully removed /tmp/{archive_filename.split('.')[0]} on {hostname}.")
                    else:
                        print(f"Error removing /tmp/{archive_filename.split('.')[0]} on {hostname}: {cleanup_result.stderr}")

                    if oem_action.strip().upper() == "OEM_RESTART":
                        reload_oem_agent(hostname)

                else:
                    print(f"Error copying {archive_filename.split('.')[0]} to {os.path.dirname(full_path)} on {hostname}: {copy_result.stderr}")

            else:
                print(f"Error extracting {archive_filename.split('.')[0]} on {hostname}: {extract_result.stderr}")
        else:
            print(f"Error transferring {archive_path} to {hostname}: {transfer_result.stderr}")

    elif action.strip().upper() == "DELETE":
        zip_directory(hostname, full_path, log_file)
        remove_directory(hostname, full_path)

def main():
    csv_file = 'backup.csv'
    log_file = get_timestamped_filename('zipped_files')  # Timestamped log file
    jdk_archive_path = 'jdk.tar.gz'
    jre_archive_path = 'jre.tar.gz'

    with open(csv_file, mode='r') as file:
        reader = csv.reader(file)
        for row in reader:
            if len(row) < 3 or len(row) > 4:
                print(f"Skipping invalid row: {row}")
                continue

            hostname, full_path, action = row[:3]
            oem_action = row[3] if len(row) == 4 else ""

            if action.strip().upper() == "JRE_UPGRADE":
                zip_and_remove_directory(hostname.strip(), full_path.strip(), log_file, jre_archive_path, action.strip(), oem_action.strip())
            else:
                zip_and_remove_directory(hostname.strip(), full_path.strip(), log_file, jdk_archive_path, action.strip(), oem_action.strip())

if __name__ == "__main__":
    main()
